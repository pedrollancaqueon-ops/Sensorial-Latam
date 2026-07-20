"""
Procesa todos los .xlsx de la raiz del proyecto.
Estructura esperada por archivo:
  - Nombre de hoja = codigo de servicio (HLDL, CLDL, etc.)
  - Fila 2: etiquetas de ciclo en columnas C, H, M, R
  - Fila 3: cabeceras por ciclo (Component, Description, Photo, GR/UN...)
  - Filas de datos:
      - Cabecera de seccion: col B tiene texto (nombre de componente)
      - Sub-filas: col B tiene numero (peso/cantidad), col C/H/M/R tiene descripcion
  - Imagenes anclan en la columna Photo de cada ciclo (D/I/N/S = cols 4/9/14/19)
Cada archivo genera su propia subcarpeta en catalog/images/<nombre_archivo>/
Todo se acumula en catalog/catalog_index.json
"""

import sys
import json
import re
from io import BytesIO
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

PROJECT_ROOT = Path(__file__).parent.parent
IMAGES_BASE  = Path(__file__).parent / "images"
INDEX_PATH   = Path(__file__).parent / "catalog_index.json"

# Columnas base (0-indexed) para cada ciclo — columnas D/I/N/S = índices 3/8/13/18
CYCLE_COLS = {
    1: {"desc": 2, "photo": 3},   # C, D
    2: {"desc": 7, "photo": 8},   # H, I
    3: {"desc": 12, "photo": 13}, # M, N
    4: {"desc": 17, "photo": 18}, # R, S
}
# Foto base índices para detección de offset
_BASE_PHOTO_COLS = [v["photo"] for v in CYCLE_COLS.values()]  # [3, 8, 13, 18]

PHOTO_COL_TO_CYCLE = {v["photo"]: k for k, v in CYCLE_COLS.items()}
COMPONENT_COL = 1  # col B (0-indexed)


def _build_photo_col_to_cycle(ws) -> dict[int, int]:
    """Detecta qué columnas contienen imágenes y construye el mapeo col→ciclo.
    Algunas hojas (p.ej. RG) tienen las fotos desplazadas +1 respecto a las LH."""
    imgs = getattr(ws, "_images", [])
    if not imgs:
        return PHOTO_COL_TO_CYCLE

    img_cols = set()
    for img in imgs:
        if hasattr(img.anchor, "_from"):
            img_cols.add(img.anchor._from.col)

    # Probar offset 0, 1, -1 ... hasta encontrar el que coincide con imágenes reales
    for offset in (0, 1, -1, 2):
        candidate = {b + offset: cycle for b, cycle in zip(_BASE_PHOTO_COLS, CYCLE_COLS.keys())}
        if img_cols & set(candidate.keys()):  # al menos una columna coincide
            return candidate

    # Fallback: mapear las columnas reales en orden a los ciclos 1,2,3,4
    sorted_cols = sorted(img_cols)[:4]
    return {col: cycle for col, cycle in zip(sorted_cols, CYCLE_COLS.keys())}


def slugify(text: str) -> str:
    text = str(text).lower().strip()
    for src, dst in [("a","a"),("a","a"),("a","a"),("e","e"),("e","e"),("e","e"),
                     ("i","i"),("i","i"),("i","i"),("o","o"),("o","o"),("o","o"),
                     ("u","u"),("u","u"),("u","u"),("n","n"),
                     ("á","a"),("à","a"),("ä","a"),("é","e"),("è","e"),("ë","e"),
                     ("í","i"),("ì","i"),("ï","i"),("ó","o"),("ò","o"),("ö","o"),
                     ("ú","u"),("ù","u"),("ü","u"),("ñ","n")]:
        text = text.replace(src, dst)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")[:40]


def is_section_header(val) -> bool:
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return False
    text = str(val).strip()
    if not text or text.lower() in ("none", "nan", "component"):
        return False
    return True


def _is_whole_quantity(val) -> bool:
    """True si el valor parece una cantidad entera (ej. 1.0, 2.0), no un factor raro (ej. 1.0512)."""
    try:
        f = float(val)
        return abs(f - round(f)) < 0.01
    except (TypeError, ValueError):
        return False


def parse_sheet(ws) -> list:
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    n_rows = len(all_rows)

    section_rows = []
    for i, row in enumerate(all_rows):
        b_val = row[COMPONENT_COL] if len(row) > COMPONENT_COL else None
        if is_section_header(b_val):
            section_rows.append((i, str(b_val).strip()))

    # Fallback: si col B no tiene secciones, buscar en col A
    # Solo toma filas donde col B es None o una cantidad entera (ej. 1.0), no factores raros
    if not section_rows:
        for i, row in enumerate(all_rows):
            a_val = row[0] if len(row) > 0 else None
            b_val = row[1] if len(row) > 1 else None
            if not is_section_header(a_val):
                continue
            if b_val is not None and not _is_whole_quantity(b_val):
                continue
            section_rows.append((i, str(a_val).strip()))

    if not section_rows:
        return []

    sections = []
    for idx, (row_i, component) in enumerate(section_rows):
        row_end = section_rows[idx + 1][0] if idx + 1 < len(section_rows) else n_rows
        descs = {c: [] for c in CYCLE_COLS}
        for r in range(row_i, row_end):
            data_row = all_rows[r]
            for cycle, cols in CYCLE_COLS.items():
                d = data_row[cols["desc"]] if len(data_row) > cols["desc"] else None
                if d and str(d).strip() and str(d).strip().lower() not in ("none", "nan", "description"):
                    descs[cycle].append(str(d).strip())
        sections.append({"row_start": row_i, "row_end": row_end, "component": component, "descs": descs})

    return sections


def find_section(sections, anchor_row, tolerance=3):
    for s in sections:
        if s["row_start"] <= anchor_row < s["row_end"]:
            return s
    for s in sections:
        if s["row_start"] - tolerance <= anchor_row < s["row_end"]:
            return s
    return None


def procesar_archivo(excel_path: Path, catalog: list) -> tuple[int, int]:
    source_name = excel_path.stem.strip()
    images_dir  = IMAGES_BASE / source_name
    images_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"Archivo: {excel_path.name}")
    print(f"Carpeta: catalog/images/{source_name}/")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    imagenes_guardadas = 0
    items_sin_imagen = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        code = sheet_name.strip()

        sections = parse_sheet(ws)
        if not sections:
            continue

        image_map = {}
        photo_col_map = _build_photo_col_to_cycle(ws)
        if hasattr(ws, "_images"):
            for img_obj in ws._images:
                anchor = img_obj.anchor
                if not hasattr(anchor, "_from"):
                    continue
                anchor_row = anchor._from.row
                anchor_col = anchor._from.col
                cycle = photo_col_map.get(anchor_col)
                if cycle is None:
                    continue
                section = find_section(sections, anchor_row)
                if section is None:
                    continue
                sec_idx = sections.index(section)
                if hasattr(img_obj, "ref") and isinstance(img_obj.ref, BytesIO):
                    img_obj.ref.seek(0)
                    img_bytes = img_obj.ref.read()
                    if img_bytes:
                        image_map[(sec_idx, cycle)] = img_bytes

        for sec_idx, section in enumerate(sections):
            for cycle in sorted(CYCLE_COLS.keys()):
                desc_list = section["descs"].get(cycle, [])
                description = ", ".join(desc_list) if desc_list else None
                img_bytes = image_map.get((sec_idx, cycle))
                img_path = None

                if img_bytes:
                    img_name = f"{code}_cycle{cycle}_{slugify(section['component'])}.jpg"
                    (images_dir / img_name).write_bytes(img_bytes)
                    img_path = f"catalog/images/{source_name}/{img_name}"
                    imagenes_guardadas += 1
                else:
                    items_sin_imagen += 1

                if description or img_bytes:
                    catalog.append({
                        "source": source_name,
                        "code": code,
                        "cycle": cycle,
                        "component": section["component"],
                        "description": description,
                        "image_path": img_path,
                    })

    img_count = sum(1 for f in images_dir.iterdir() if f.suffix == ".jpg")
    print(f"  Hojas: {len(wb.sheetnames)} | Imagenes: {imagenes_guardadas} | Sin imagen: {items_sin_imagen}")
    return imagenes_guardadas, items_sin_imagen


def _sort_key(p: Path):
    """Archivos con prefijo de fecha (DD.MM.YYYY-) se procesan AL FINAL para que sus
    imágenes tengan prioridad en catalog.py (el último procesado gana con >= en get_catalog_images)."""
    if re.match(r"^\d{2}\.\d{2}\.\d{4}", p.stem):
        return (1, p.stem)
    return (0, p.stem)


def procesar_todos():
    xlsx_files = sorted(PROJECT_ROOT.glob("*.xlsx"), key=_sort_key)
    if not xlsx_files:
        print("No se encontraron archivos .xlsx en la raiz del proyecto.")
        return

    print(f"Archivos encontrados: {len(xlsx_files)}")
    for f in xlsx_files:
        print(f"  - {f.name}")

    catalog = []
    total_imgs = 0
    total_sin  = 0

    for excel_path in xlsx_files:
        imgs, sin = procesar_archivo(excel_path, catalog)
        total_imgs += imgs
        total_sin  += sin

    INDEX_PATH.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n{'='*60}")
    print(f"TOTAL archivos    : {len(xlsx_files)}")
    print(f"TOTAL items       : {len(catalog)}")
    print(f"TOTAL con imagen  : {total_imgs}")
    print(f"TOTAL sin imagen  : {total_sin}")
    print(f"\nIndice: {INDEX_PATH}")


if __name__ == "__main__":
    procesar_todos()
